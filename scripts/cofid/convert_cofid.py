"""
Convert CoFID 2021 (monosaccharide-equivalent carbohydrate) into UK FIC
by-weight nutrition values, per 100 g, for the seven mandatory nutrients + fibre.

Conversion factors (from the CoFID user guide + basic carbohydrate chemistry):
  - monosaccharides (glucose, galactose, fructose): already by weight     → ×1.00
  - disaccharides (sucrose, maltose, lactose): monosacc-equiv → by weight → ×0.95 (342/360)
  - starch & oligosaccharides (polysaccharides): → by weight              → ×0.90 (162/180)
Energy is RECOMPUTED with FIC Annex XIV factors (never copied from CoFID):
  kcal = 4·protein + 9·fat + 4·carb_byweight + 2·fibre
  kJ   = 17·protein + 37·fat + 17·carb_byweight + 8·fibre
Salt = sodium(mg)/1000 × 2.5.
"""
import openpyxl, json, re

MONO, DI, POLY = 1.00, 0.95, 0.90
SRC = "cofid2021.xlsx"

def num(v):
    """CoFID cells: numbers, 'N' (not analysed), 'Tr'/'trace', '' , '(1.2)' estimated, 'N/A'."""
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s in ("", "N", "n", "N/A"): return None
    if s.lower() in ("tr", "trace"): return 0.0
    s = s.strip("()")            # estimated values in brackets → use the number
    s = s.replace("*", "").strip()
    try: return float(s)
    except ValueError: return None

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# ── Proximates: header row is row 1; two label rows (2,3) then data ───────────
prox = wb["1.3 Proximates"]
rows = list(prox.iter_rows(values_only=True))
# column indices (verified from inspection)
C = dict(code=0, name=1, desc=2, group=3, prot=9, fat=10, cho=11,
         starch=14, oligo=15, tot_sug=16, glucose=17, galactose=18,
         fructose=19, sucrose=20, maltose=21, lactose=22, alcohol=23,
         nsp=24, aoac=25)

# ── Inorganics: sodium for salt ──────────────────────────────────────────────
inorg = list(wb["1.4 Inorganics"].iter_rows(values_only=True))
sodium_by_code = {}
for r in inorg[3:]:
    if r[0]: sodium_by_code[str(r[0]).strip()] = num(r[7])  # col 7 = Sodium (mg)

out = []
flags = {"sugar_breakdown_missing": 0, "fibre_nsp_fallback": 0, "no_fibre": 0, "carb_from_total_only": 0}

for r in rows[3:]:
    code = r[C["code"]]
    if not code or not str(code).strip(): continue
    code = str(code).strip()
    name = str(r[C["name"]]).strip() if r[C["name"]] else ""
    if not name: continue

    protein = num(r[C["prot"]]) or 0.0
    fat     = num(r[C["fat"]]) or 0.0
    cho_m   = num(r[C["cho"]])          # carbohydrate, monosaccharide equiv
    starch_m= num(r[C["starch"]])
    oligo_m = num(r[C["oligo"]])
    totsug_m= num(r[C["tot_sug"]])
    glu, gal, fru = num(r[C["glucose"]]), num(r[C["galactose"]]), num(r[C["fructose"]])
    suc, mal, lac = num(r[C["sucrose"]]), num(r[C["maltose"]]), num(r[C["lactose"]])
    aoac, nsp = num(r[C["aoac"]]), num(r[C["nsp"]])

    # ── Sugars → by weight ───────────────────────────────────────────────────
    has_breakdown = all(x is not None for x in (glu, gal, fru, suc, mal, lac))
    if has_breakdown:
        sugars_bw = (glu + gal + fru) * MONO + (suc + mal + lac) * DI
    elif totsug_m is not None:
        # No per-sugar split: total sugars is monosacc-equiv. Apply the
        # disaccharide factor as a conservative default (≤5% error, within
        # FIC tolerance) and flag it.
        sugars_bw = totsug_m * DI
        flags["sugar_breakdown_missing"] += 1
    else:
        sugars_bw = None

    # ── Carbohydrate → by weight ─────────────────────────────────────────────
    # Anchor to CoFID's authoritative total CHO (monosacc) and scale it by a
    # per-food factor derived from its starch/sugar composition, bounded to
    # [POLY, MONO]. Anchoring (vs summing components) keeps by-weight physically
    # valid — never above the monosacc total, never below 0.90× it — even when
    # the component columns don't perfectly reconcile with the total.
    if cho_m is not None and starch_m is not None and totsug_m is not None and (starch_m + totsug_m) > 0:
        oligo_res = oligo_m if oligo_m is not None else max(0.0, cho_m - starch_m - totsug_m)
        sugar_factor = (sugars_bw / totsug_m) if (sugars_bw is not None and totsug_m > 0) else DI
        denom = starch_m + totsug_m + oligo_res
        factor = (POLY*starch_m + sugar_factor*totsug_m + POLY*oligo_res) / denom
        factor = max(POLY, min(MONO, factor))
        carb_bw = cho_m * factor
    elif cho_m is not None:
        # No usable breakdown — blended fallback (midpoint of 0.90–1.00).
        carb_bw = cho_m * 0.94
        flags["carb_from_total_only"] += 1
    else:
        carb_bw = None

    # ── Saturates (per 100 g food, already by weight) ────────────────────────
    sat = num(r[27])  # Satd FA /100g fd

    # ── Fibre (AOAC preferred; NSP fallback) ─────────────────────────────────
    if aoac is not None:
        fibre = aoac
    elif nsp is not None:
        fibre = nsp; flags["fibre_nsp_fallback"] += 1
    else:
        fibre = None; flags["no_fibre"] += 1

    # ── Salt ─────────────────────────────────────────────────────────────────
    na_mg = sodium_by_code.get(code)
    salt = (na_mg / 1000.0) * 2.5 if na_mg is not None else None

    # ── Energy (FIC Annex XIV, recomputed) ───────────────────────────────────
    c = carb_bw or 0.0; fb = fibre or 0.0
    alc = num(r[C["alcohol"]]) or 0.0
    kcal = 4*protein + 9*fat + 4*c + 2*fb + 7*alc
    kj   = 17*protein + 37*fat + 17*c + 8*fb + 29*alc

    out.append({
        "code": code, "name": name,
        "group": str(r[C["group"]]).strip() if r[C["group"]] else "",
        "kcal": round(kcal, 1), "kj": round(kj),
        "fat": round(fat, 2), "saturates": round(sat, 2) if sat is not None else None,
        "carbohydrate": round(carb_bw, 2) if carb_bw is not None else None,
        "sugars": round(sugars_bw, 2) if sugars_bw is not None else None,
        "fibre": round(fibre, 2) if fibre is not None else None,
        "protein": round(protein, 2), "salt": round(salt, 3) if salt is not None else None,
    })

json.dump(out, open("../../lib/nutrition/cofid.json", "w"), separators=(",", ":"))
print(f"Converted {len(out)} foods → ../../lib/nutrition/cofid.json")
print("Flags:", flags)
